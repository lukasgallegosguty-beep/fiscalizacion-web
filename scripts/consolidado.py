#!/usr/bin/env python3
"""
Consolidado mensual de la fiscalización web de DM (ISP/ANDIM).

Cierra el mes: junta en un solo Excel los casos del mes que YA pasaron por un
inspector y quedaron en pie, para que los tres decidan en la reunión de la
semana 4 cuáles se procesan como denuncia.

Qué entra (y de dónde):

  A. Hallazgos marcados NO REGISTRADO por la rutina y confirmados por el
     inspector con un "Correcto" en la columna «Decisión final».
  B. URLs que el inspector escribió a mano en «Observaciones del inspector» de
     la hoja de marketplace. Son productos que él encontró navegando y que la
     rutina no vio (Mercado Libre devuelve 403 a la rutina). Entran SIN cruce
     automático, marcados «POR VERIFICAR»: nadie los comparó todavía contra el
     listado ISP.

Qué NO entra como caso denunciable: los hallazgos que la rutina dio por
REGISTRADO y el inspector marcó "Incorrecto". Es tentador leerlos como
infracciones que la rutina dejó pasar, pero al revisar los cinco de agosto el
inspector no estaba diciendo eso: decía "el enlace arroja Error - 404" y "no es
posible confirmar mediante imágenes si el producto cuenta con registro
vigente". Son verificaciones que no se pudieron completar, no productos
ilegales. Llevarlos a la mesa de denuncias sería acusar por un enlace caído, así
que van a la hoja «Discrepancias sin resolver», sin columna de denuncia.

Qué NO entra: lo que sigue sin revisar. Un reporte que el inspector aún no
devolvió no aporta casos, y eso se dice explícitamente en la hoja «Cobertura
del mes» y en el resumen — nunca se omite en silencio.

Uso:
    python3 scripts/consolidado.py --mes 2026-09
    python3 scripts/consolidado.py --json
"""

import argparse
import json
import os
import re
import sys
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))

import rotacion  # noqa: E402
from feedback import veredicto  # noqa: E402

AZUL_ISP = "003366"
FILA_ALT = "F2F6FA"
ROJO = "F8D7D7"
AMARILLO = "FFF6CC"
GRIS = "EDEDED"
NARANJA = "FCE4D0"
BORDE = Border(*[Side(style="thin", color="C8CDD4")] * 4)

ORIGEN_CONFIRMADO = "Hallazgo NO REGISTRADO confirmado por el inspector"
ORIGEN_DISCREPANCIA = "La rutina lo dio por REGISTRADO y el inspector marcó Incorrecto"
ORIGEN_MANUAL = "Detectado por el inspector en revisión manual del marketplace"

COLUMNAS = [
    ("Categoría", 30, "categoria"),
    ("Fecha de detección", 14, "fecha"),
    ("Nombre de DM ofertado", 34, "nombre_dm"),
    ("URL", 46, "url"),
    ("Título de la publicación", 36, "titulo"),
    ("Oferente", 24, "oferente"),
    ("Coincidencia", 12, "coincidencia"),
    ("Nombre del producto con el que coincide", 30, "producto_isp"),
    ("Registro del producto con el que coincide", 20, "registro_isp"),
    ("Clasificación", 18, "clasificacion"),
    ("Observaciones", 56, "observaciones"),
    ("Origen del caso", 34, "origen"),
    ("Inspector que revisó", 22, "inspector"),
    ("Observaciones del inspector", 56, "obs_inspector"),
    ("¿Se procesa como denuncia?", 22, "denuncia"),
    ("Justificación de la decisión", 60, "justificacion"),
]

COLUMNAS_DISCREPANCIA = [
    ("Categoría", 28, "categoria"),
    ("Fecha", 12, "fecha"),
    ("Nombre de DM ofertado", 36, "nombre_dm"),
    ("URL", 46, "url"),
    ("Registro que citó la rutina", 22, "registro_isp"),
    ("Decisión del inspector", 18, "decision"),
    ("Por qué lo marcó así", 70, "obs_inspector"),
]

COLUMNAS_NOTAS = [
    ("Categoría", 28, "categoria"),
    ("Fecha", 12, "fecha"),
    ("Marketplace / ámbito", 34, "marketplace"),
    ("Nota del inspector", 100, "texto"),
]

COLUMNAS_COBERTURA = [
    ("Fecha", 12, "fecha"),
    ("Semana del mes", 14, "semana"),
    ("Categoría", 32, "categoria"),
    ("Inspector asignado", 24, "inspector"),
    ("Reporte emitido", 14, "emitido"),
    ("Revisado y devuelto", 18, "revisado"),
    ("Casos que aportó al consolidado", 20, "casos"),
    ("Estado", 60, "estado"),
]

RE_URL = re.compile(r"https?://[^\s,;<>\"')\]]+")


def _limpiar_url(u):
    """Quita el fragmento de tracking y la puntuación final.

    Las URLs que pegan los inspectores vienen de la búsqueda de Mercado Libre y
    arrastran 300 caracteres de `#polycard_client=...&position=53&tracking_id=...`
    que no identifican el producto y solo estorban en una denuncia.
    """
    u = u.split("#")[0].rstrip(".,;)")
    return u


def _clave_url(u):
    return _limpiar_url(str(u or "")).strip().lower().rstrip("/")


def _txt(v):
    return "" if v is None else str(v).strip()


def _indice(ws):
    idx = {}
    for celda in ws[1]:
        if celda.value:
            idx[rotacion._sin_tildes(str(celda.value).strip())] = celda.column
    return idx


def _inspector_de_archivo(info):
    """Quién revisó ese reporte: primero el historial, después el calendario.

    Los reportes de agosto se generaron con el esquema anterior y su semana ya
    no tiene inspector asignado bajo el calendario mensual. Antes que inventar
    uno, se deja constancia de que no se pudo determinar.
    """
    f = info["fecha"]
    if f:
        for h in rotacion.leer_historial():
            if h.get("categoria") == info["slug"] and h.get("fecha") == f.isoformat():
                if h.get("inspector"):
                    return h["inspector"]
        insp = rotacion.inspector_de(f)
        if insp:
            return insp["email"]
    return "(no determinado)"


def _leer(info):
    """Extrae de un Excel revisado los casos que van al consolidado."""
    cat = rotacion.POR_SLUG[info["slug"]]
    fecha = info["fecha"].strftime("%d-%m-%Y") if info["fecha"] else ""
    quien = _inspector_de_archivo(info)
    casos, descartes, mkt_ignorado, discrepancias = [], [], [], []

    wb = load_workbook(info["ruta"], read_only=True, data_only=True)

    if "Hallazgos" in wb.sheetnames:
        ws = wb["Hallazgos"]
        idx = _indice(ws)

        def col(nombre):
            return idx.get(nombre)

        campos = {
            "nombre_dm": col("nombre de dm ofertado"), "url": col("url"),
            "titulo": col("titulo de la publicacion"), "oferente": col("oferente"),
            "coincidencia": col("coincidencia"),
            "producto_isp": col("nombre del producto con el que coincide"),
            "registro_isp": col("registro del producto con el que coincide"),
            "clasificacion": col("clasificacion"), "observaciones": col("observaciones"),
            "decision": col("decision final"), "obs": col("observaciones del inspector"),
        }
        for fila in ws.iter_rows(min_row=2):
            def v(c):
                return _txt(fila[c - 1].value) if c and len(fila) >= c else ""

            clasif = v(campos["clasificacion"]).upper()
            ver = veredicto(v(campos["decision"]))

            if clasif == "REGISTRADO" and ver == "INCORRECTO":
                # El inspector contradijo a la rutina, pero eso no basta para
                # denunciar: en agosto los cinco casos así eran enlaces caídos o
                # publicaciones sin imágenes suficientes para verificar. Van a
                # una hoja aparte, sin columna de denuncia.
                discrepancias.append({
                    "categoria": cat["nombre"], "fecha": fecha,
                    "nombre_dm": v(campos["nombre_dm"]),
                    "url": _limpiar_url(v(campos["url"])),
                    "registro_isp": v(campos["registro_isp"]),
                    "decision": v(campos["decision"]),
                    "obs_inspector": v(campos["obs"]),
                })
                continue

            if clasif == "NO REGISTRADO" and ver == "CORRECTO":
                origen = ORIGEN_CONFIRMADO
            else:
                if clasif in ("NO REGISTRADO", "REGISTRADO"):
                    descartes.append({
                        "nombre": v(campos["nombre_dm"]), "clasificacion": clasif,
                        "decision": v(campos["decision"]) or "(sin decisión)",
                        "veredicto": ver or "SIN DECIDIR",
                    })
                continue

            casos.append({
                "categoria": cat["nombre"], "fecha": fecha,
                "nombre_dm": v(campos["nombre_dm"]),
                "url": _limpiar_url(v(campos["url"])),
                "titulo": v(campos["titulo"]), "oferente": v(campos["oferente"]),
                "coincidencia": v(campos["coincidencia"]),
                "producto_isp": v(campos["producto_isp"]),
                "registro_isp": v(campos["registro_isp"]),
                "clasificacion": "NO REGISTRADO",
                "observaciones": v(campos["observaciones"]),
                "origen": origen, "inspector": quien,
                "obs_inspector": v(campos["obs"]),
                "denuncia": "", "justificacion": "",
                "_slug": info["slug"], "_archivo": info["archivo"],
            })

    for hoja in wb.sheetnames:
        if not hoja.startswith("Búsquedas"):
            continue
        ws = wb[hoja]
        idx = _indice(ws)
        c_obs, c_mkt = idx.get("observaciones del inspector"), idx.get("marketplace")
        for fila in ws.iter_rows(min_row=2):
            def v(c):
                return _txt(fila[c - 1].value) if c and len(fila) >= c else ""

            obs, donde = v(c_obs), v(c_mkt)
            if not obs:
                continue
            urls = [_limpiar_url(u) for u in RE_URL.findall(obs)]
            if not urls:
                # Comentario sin enlaces: no hay producto que tabular. Se
                # arrastra al resumen para que nadie lo dé por perdido.
                mkt_ignorado.append({
                    "categoria": cat["nombre"], "fecha": fecha,
                    "marketplace": donde, "texto": obs,
                })
                continue
            for u in urls:
                casos.append({
                    "categoria": cat["nombre"], "fecha": fecha,
                    "nombre_dm": "(no consignado por el inspector)",
                    "url": u, "titulo": "", "oferente": donde,
                    "coincidencia": "", "producto_isp": "", "registro_isp": "",
                    "clasificacion": "POR VERIFICAR",
                    "observaciones": (
                        "Detección manual del inspector sobre "
                        f"{donde or 'el marketplace'}. No pasó por el cruce "
                        "automático contra el listado ISP: verificar registro "
                        "antes de resolver la denuncia."
                    ),
                    "origen": ORIGEN_MANUAL, "inspector": quien,
                    "obs_inspector": obs,
                    "denuncia": "", "justificacion": "",
                    "_slug": info["slug"], "_archivo": info["archivo"],
                })
    wb.close()
    return casos, descartes, mkt_ignorado, discrepancias


def recopilar(anio, mes):
    """Junta los casos de todos los reportes revisados con fecha en ese mes."""
    revisados, no_atribuidos = rotacion.archivos_revisados(anio=anio, mes=mes)

    casos, descartes, mkt_ignorado, discrepancias, por_archivo = [], [], [], [], {}
    for info in revisados:
        c, d, m, x = _leer(info)
        por_archivo[info["archivo"]] = len(c)
        casos.extend(c)
        descartes.extend(d)
        mkt_ignorado.extend(m)
        discrepancias.extend(x)

    # Una misma oferta puede aparecer en dos reportes del mes. Gana la primera
    # (la más antigua), que es la que trae el cruce contra el listado ISP.
    vistas, unicos, duplicados = set(), [], 0
    for c in sorted(casos, key=lambda x: (x["clasificacion"] == "POR VERIFICAR", x["fecha"])):
        k = _clave_url(c["url"])
        if k and k in vistas:
            duplicados += 1
            continue
        if k:
            vistas.add(k)
        unicos.append(c)

    unicos.sort(key=lambda c: (c["categoria"], c["clasificacion"] == "POR VERIFICAR", c["fecha"]))
    return {
        "casos": unicos,
        "revisados": revisados,
        "no_atribuidos": no_atribuidos,
        "descartes": descartes,
        "discrepancias": discrepancias,
        "mkt_sin_enlaces": mkt_ignorado,
        "duplicados": duplicados,
        "casos_por_archivo": por_archivo,
    }


def _encabezados(ws, columnas):
    fuente = Font(bold=True, color="FFFFFF", size=11)
    relleno = PatternFill("solid", fgColor=AZUL_ISP)
    for i, (titulo, ancho, _) in enumerate(columnas, 1):
        c = ws.cell(row=1, column=i, value=titulo)
        c.font, c.fill, c.border = fuente, relleno, BORDE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"


def _escribir(ws, columnas, filas):
    for n, item in enumerate(filas, start=2):
        for i, (_, _, clave) in enumerate(columnas, 1):
            c = ws.cell(row=n, column=i, value=item.get(clave, ""))
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDE
            if n % 2 == 0:
                c.fill = PatternFill("solid", fgColor=FILA_ALT)


def _nota(ws, fila, texto):
    ws.cell(row=fila, column=1, value=texto).font = Font(italic=True, size=9)


def generar(anio, mes, datos, plan, salida):
    casos = datos["casos"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Casos del mes"
    _encabezados(ws, COLUMNAS)

    filas = casos or [{
        "categoria": "", "fecha": "", "nombre_dm": "Sin casos confirmados",
        "url": "", "titulo": "", "oferente": "", "coincidencia": "",
        "producto_isp": "", "registro_isp": "", "clasificacion": "SIN CASOS",
        "observaciones": (
            f"Ningún reporte revisado de {rotacion.MESES[mes - 1]} {anio} dejó "
            "hallazgos NO REGISTRADO confirmados ni detecciones manuales con "
            "enlace. Ver la hoja «Cobertura del mes»: puede deberse a que los "
            "reportes aún no fueron devueltos por los inspectores."
        ),
        "origen": "", "inspector": "", "obs_inspector": "",
        "denuncia": "", "justificacion": "",
    }]

    i_clasif = next(i for i, c in enumerate(COLUMNAS, 1) if c[2] == "clasificacion")
    i_den = next(i for i, c in enumerate(COLUMNAS, 1) if c[2] == "denuncia")
    i_jus = next(i for i, c in enumerate(COLUMNAS, 1) if c[2] == "justificacion")

    for n, item in enumerate(filas, start=2):
        for i, (_, _, clave) in enumerate(COLUMNAS, 1):
            c = ws.cell(row=n, column=i, value=item.get(clave, ""))
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDE
            if n % 2 == 0:
                c.fill = PatternFill("solid", fgColor=FILA_ALT)
        cl = ws.cell(row=n, column=i_clasif)
        if item.get("clasificacion") == "NO REGISTRADO":
            cl.fill = PatternFill("solid", fgColor=ROJO)
            cl.font = Font(bold=True, color="A11B1B")
        elif item.get("clasificacion") == "POR VERIFICAR":
            cl.fill = PatternFill("solid", fgColor=NARANJA)
            cl.font = Font(bold=True, color="8A4B12")
        else:
            cl.fill = PatternFill("solid", fgColor=GRIS)
        for i in (i_den, i_jus):
            ws.cell(row=n, column=i).fill = PatternFill("solid", fgColor=AMARILLO)

    ultima = len(filas) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{ultima}"

    if casos:
        dv = DataValidation(type="list", formula1='"SÍ,NO"', allow_blank=True,
                            showDropDown=False, showErrorMessage=True,
                            errorTitle="Valor no válido",
                            error="Responde SÍ o NO: ¿se procesa este caso como denuncia?")
        ws.add_data_validation(dv)
        letra = get_column_letter(i_den)
        dv.add(f"{letra}2:{letra}{ultima}")

    # --- Discrepancias: el inspector contradijo a la rutina, pero eso NO es
    # una infracción confirmada. Sin columna de denuncia, a propósito.
    if datos["discrepancias"]:
        ws3 = wb.create_sheet("Discrepancias sin resolver")
        _encabezados(ws3, COLUMNAS_DISCREPANCIA)
        _escribir(ws3, COLUMNAS_DISCREPANCIA, datos["discrepancias"])
        _nota(ws3, len(datos["discrepancias"]) + 3, (
            f"Estos {len(datos['discrepancias'])} casos los clasificó la rutina como REGISTRADO y el "
            "inspector marcó «Incorrecto». NO están en la hoja de casos y no se proponen como denuncia: "
            "en la práctica el motivo suele ser un enlace caído o una publicación sin imágenes "
            "suficientes para verificar, no un producto sin registro. Si en la reunión se determina que "
            "alguno sí es una infracción, se pasa a mano a la hoja «Casos del mes»."
        ))

    # --- Notas de los inspectores sin enlace: no tabulables, pero tampoco se
    # tiran. Muchas explican por qué un marketplace no se pudo revisar.
    if datos["mkt_sin_enlaces"]:
        ws4 = wb.create_sheet("Notas de marketplace")
        _encabezados(ws4, COLUMNAS_NOTAS)
        _escribir(ws4, COLUMNAS_NOTAS, datos["mkt_sin_enlaces"])
        _nota(ws4, len(datos["mkt_sin_enlaces"]) + 3, (
            "Observaciones que los inspectores escribieron en la hoja de marketplace y que no traen "
            "enlaces, por lo que no hay producto que tabular. Se conservan porque explican qué se "
            "revisó a mano y qué quedó bloqueado."
        ))

    # --- Cobertura ---
    ws2 = wb.create_sheet("Cobertura del mes")
    _encabezados(ws2, COLUMNAS_COBERTURA)
    por_archivo = datos["casos_por_archivo"]
    atribuido = {i["archivo"]: i for i in datos["revisados"]}
    devuelto_por_slug = {}
    for a, info in atribuido.items():
        devuelto_por_slug.setdefault((info["slug"], info["fecha"]), a)

    filas_cob = []
    for b in plan["bloques_esperados"]:
        f = date.fromisoformat(b["fecha"])
        arch = devuelto_por_slug.get((b["slug"], f))
        casos_n = por_archivo.get(arch, 0) if arch else 0
        if not b["reporte"]:
            estado = "La corrida no llegó a generar el reporte."
        elif not arch:
            estado = "Reporte emitido pero AÚN NO devuelto por el inspector: no aporta casos a este consolidado."
        elif casos_n:
            estado = f"Revisado. Aportó {casos_n} caso(s)."
        else:
            estado = "Revisado. Sin casos confirmados que consolidar."
        filas_cob.append({
            "fecha": b["fecha_dmy"], "semana": f"Semana {b['semana_mes']}",
            "categoria": b["categoria"], "inspector": b["inspector"],
            "emitido": "Sí" if b["reporte"] else "No",
            "revisado": "Sí" if arch else "No",
            "casos": casos_n, "estado": estado,
        })

    # Reportes devueltos que no estaban en el calendario del mes (p. ej. los del
    # esquema anterior). Se listan igual: aportaron casos y hay que poder
    # rastrear de dónde salió cada fila.
    esperados = {(b["slug"], date.fromisoformat(b["fecha"])) for b in plan["bloques_esperados"]}
    for info in datos["revisados"]:
        if (info["slug"], info["fecha"]) in esperados:
            continue
        filas_cob.append({
            "fecha": info["fecha"].strftime("%d-%m-%Y") if info["fecha"] else "",
            "semana": "fuera de calendario",
            "categoria": rotacion.POR_SLUG[info["slug"]]["nombre"],
            "inspector": _inspector_de_archivo(info),
            "emitido": "Sí", "revisado": "Sí",
            "casos": por_archivo.get(info["archivo"], 0),
            "estado": f"Devuelto fuera del calendario mensual ({info['archivo']}).",
        })

    for n, item in enumerate(filas_cob, start=2):
        for i, (_, _, clave) in enumerate(COLUMNAS_COBERTURA, 1):
            c = ws2.cell(row=n, column=i, value=item.get(clave, ""))
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDE
            if n % 2 == 0:
                c.fill = PatternFill("solid", fgColor=FILA_ALT)
        if item["revisado"] == "No":
            ws2.cell(row=n, column=6).fill = PatternFill("solid", fgColor=NARANJA)

    os.makedirs(os.path.dirname(os.path.abspath(salida)), exist_ok=True)
    wb.save(salida)
    return salida


def resumir(anio, mes, datos, plan, salida):
    casos = datos["casos"]
    def cuenta(origen):
        return sum(1 for c in casos if c["origen"] == origen)
    por_cat = {}
    for c in casos:
        por_cat[c["categoria"]] = por_cat.get(c["categoria"], 0) + 1
    return {
        "periodo": f"{anio}-{mes:02d}",
        "mes": rotacion.MESES[mes - 1],
        "anio": anio,
        "archivo": salida,
        "total_casos": len(casos),
        "confirmados": cuenta(ORIGEN_CONFIRMADO),
        "deteccion_manual_marketplace": cuenta(ORIGEN_MANUAL),
        "discrepancias_sin_resolver": len(datos["discrepancias"]),
        "por_categoria": por_cat,
        "archivos_revisados": [i["archivo"] for i in datos["revisados"]],
        "reportes_emitidos": plan["reportes_emitidos"],
        "reportes_esperados": len(plan["bloques_esperados"]),
        "pendientes_de_revision": plan["sin_revisar"],
        "no_emitidos": plan["sin_emitir"],
        "urls_duplicadas_omitidas": datos["duplicados"],
        "archivos_no_atribuidos": datos["no_atribuidos"],
        "decisiones_sin_tomar": sum(1 for d in datos["descartes"] if d["veredicto"] == "SIN DECIDIR"),
        "observaciones_marketplace_sin_enlace": len(datos["mkt_sin_enlaces"]),
        "destinatarios": plan["destinatarios"],
        "reunion": plan["reunion_inicio"],
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--mes", help="periodo YYYY-MM (por defecto, el mes de la semana en curso)")
    ap.add_argument("--salida", help="ruta del .xlsx (por defecto, la del plan)")
    ap.add_argument("--json", action="store_true", help="salida JSON")
    args = ap.parse_args()

    if args.mes:
        anio, mes = (int(x) for x in args.mes.split("-"))
        plan = rotacion.plan_consolidacion(rotacion.lunes_cierre(anio, mes))
    else:
        plan = rotacion.plan_consolidacion()
        anio, mes = plan["anio"], plan["mes"]

    datos = recopilar(anio, mes)
    salida = args.salida or plan["archivo_salida"]
    generar(anio, mes, datos, plan, salida)
    resumen = resumir(anio, mes, datos, plan, salida)

    if args.json:
        print(json.dumps(resumen, ensure_ascii=False, indent=2))
        return 0

    print(f"Consolidado {resumen['mes']} {anio}  ->  {salida}")
    print(f"  Casos totales                    : {resumen['total_casos']}")
    print(f"    NO REGISTRADO confirmados      : {resumen['confirmados']}")
    print(f"    Detección manual (marketplace) : {resumen['deteccion_manual_marketplace']}")
    print(f"  Discrepancias sin resolver (hoja aparte, NO denunciables): "
          f"{resumen['discrepancias_sin_resolver']}")
    print(f"  Reportes del mes                 : {resumen['reportes_emitidos']}/{resumen['reportes_esperados']} emitidos, "
          f"{len(resumen['archivos_revisados'])} devueltos por inspectores")
    if resumen["pendientes_de_revision"]:
        print(f"  PENDIENTES de revisión ({len(resumen['pendientes_de_revision'])}) — no aportan casos:")
        for a in resumen["pendientes_de_revision"]:
            print(f"    - {a}")
    if resumen["decisiones_sin_tomar"]:
        print(f"  Filas sin «Decisión final» en los archivos devueltos: {resumen['decisiones_sin_tomar']}")
    if resumen["urls_duplicadas_omitidas"]:
        print(f"  URLs repetidas omitidas          : {resumen['urls_duplicadas_omitidas']}")
    if resumen["archivos_no_atribuidos"]:
        print(f"  !! Archivos en revision/ que no se pudieron atribuir: {resumen['archivos_no_atribuidos']}")
    for cat, n in sorted(resumen["por_categoria"].items(), key=lambda x: -x[1]):
        print(f"    {n:>3}  {cat}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
