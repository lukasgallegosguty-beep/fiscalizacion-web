#!/usr/bin/env python3
"""
Generador del Excel de hallazgos de fiscalización web (formato estándar ANDIM).

Toma un JSON con los hallazgos de la jornada y produce el .xlsx con el formato
que define la skill: 10 columnas fijas, hoja secundaria "Búsquedas Marketplace"
y fila "Sin hallazgos" cuando la búsqueda no arroja resultados.

Uso:
    python3 scripts/generar_reporte.py --entrada hallazgos.json --salida ruta.xlsx
    python3 scripts/generar_reporte.py --entrada hallazgos.json --auto

Estructura del JSON de entrada (todas las claves son opcionales salvo `categoria`):

    {
      "categoria": "Guantes quirúrgicos de látex",
      "fecha": "19-08-2026",
      "hallazgos": [
        {
          "nombre_dm": "...",           "url": "https://...",
          "titulo": "...",              "oferente": "...",
          "coincidencia": "NO",         "producto_isp": "",
          "registro_isp": "",           "clasificacion": "NO REGISTRADO",
          "observaciones": "..."
        }
      ],
      "marketplace": [
        {
          "marketplace": "Mercado Libre Chile", "url_busqueda": "https://...",
          "palabras_clave": "...",              "marcas_detectadas": "...",
          "cantidad_aprox": "~120",             "observaciones": "..."
        }
      ]
    }
"""

import argparse
import json
import os
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Tope de hallazgos por reporte. Es un límite de CARGA DE REVISIÓN para el
# inspector, no un límite de esfuerzo de búsqueda: la búsqueda sigue siendo
# exhaustiva y lo que excede el tope queda documentado en la hoja "Anexo" para
# que reaparezca en la corrida siguiente de la categoría.
MAX_HALLAZGOS = 10

# Paleta institucional: azul ISP para encabezados.
AZUL_ISP = "003366"
FILA_ALT = "F2F6FA"
VERDE = "D7F0D7"   # REGISTRADO
ROJO = "F8D7D7"    # NO REGISTRADO
AMARILLO = "FFF6CC"  # "Decisión final": requiere input humano
GRIS = "EDEDED"

COLUMNAS = [
    ("Nombre de DM ofertado", 34, "nombre_dm"),
    ("URL", 46, "url"),
    ("Título de la publicación", 40, "titulo"),
    ("Oferente", 26, "oferente"),
    ("Coincidencia", 14, "coincidencia"),
    ("Nombre del producto con el que coincide", 34, "producto_isp"),
    ("Registro del producto con el que coincide", 22, "registro_isp"),
    ("Clasificación", 18, "clasificacion"),
    ("Observaciones", 60, "observaciones"),
    ("Decisión final", 22, "decision_final"),
    ("Observaciones del inspector", 60, "obs_inspector"),
]

COLUMNAS_MKT = [
    ("Marketplace", 24, "marketplace"),
    ("URL de búsqueda", 52, "url_busqueda"),
    ("Palabras clave usadas", 34, "palabras_clave"),
    ("Marcas detectadas en filtros", 40, "marcas_detectadas"),
    ("Cantidad aprox. de publicaciones", 18, "cantidad_aprox"),
    ("Observaciones para el fiscalizador", 70, "observaciones"),
    ("Observaciones del inspector", 60, "obs_inspector"),
]

COLUMNAS_ANEXO = [
    ("Nombre de DM ofertado", 40, "nombre_dm"),
    ("URL", 52, "url"),
    ("Oferente", 26, "oferente"),
    ("Clasificación", 18, "clasificacion"),
    ("Observaciones", 70, "observaciones"),
]

BORDE = Border(*[Side(style="thin", color="C8CDD4")] * 4)

# Alcance de la regulación de guantes: la INTERSECCIÓN de uso médico
# (examinación o quirúrgico) y material látex/caucho. Todo lo demás queda fuera,
# tenga o no registro. Ver SKILL.md. Los dos lados fallaron en producción: 13
# acusaciones a nitrilo y vinilo el 25-08-2026, y dos a guantes negros de látex
# de uso no médico el 01-09-2026.
MATERIALES_FUERA_ALCANCE = (
    "nitrilo", "nitrile", "vinilo", "vinyl", "neopreno", "neoprene",
    "polietileno", "polyethylene", "pvc",
)
USO_MEDICO = ("examin", "quirurg", "cirug", "surgical", "exam ", "medical exam")


def _plano(texto):
    return (
        str(texto or "").lower()
        .replace("á", "a").replace("é", "e").replace("í", "i")
        .replace("ó", "o").replace("ú", "u").replace("ñ", "n")
    )


def _encabezados(ws, columnas):
    fuente = Font(bold=True, color="FFFFFF", size=11)
    relleno = PatternFill("solid", fgColor=AZUL_ISP)
    for i, (titulo, ancho, _) in enumerate(columnas, 1):
        celda = ws.cell(row=1, column=i, value=titulo)
        celda.font = fuente
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = BORDE
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def _escribir_filas(ws, columnas, filas, colorear_clasificacion=True):
    for n, item in enumerate(filas, start=2):
        alterna = (n % 2 == 0)
        for i, (_, _, clave) in enumerate(columnas, 1):
            valor = item.get(clave, "")
            celda = ws.cell(row=n, column=i, value=valor)
            celda.alignment = Alignment(vertical="top", wrap_text=True)
            celda.border = BORDE
            if alterna:
                celda.fill = PatternFill("solid", fgColor=FILA_ALT)

        if colorear_clasificacion:
            clasif = str(item.get("clasificacion", "")).strip().upper()
            col_clasif = next(i for i, c in enumerate(columnas, 1) if c[2] == "clasificacion")
            celda = ws.cell(row=n, column=col_clasif)
            if clasif == "REGISTRADO":
                celda.fill = PatternFill("solid", fgColor=VERDE)
                celda.font = Font(bold=True, color="1E6B1E")
            elif clasif == "NO REGISTRADO":
                celda.fill = PatternFill("solid", fgColor=ROJO)
                celda.font = Font(bold=True, color="A11B1B")
            else:
                celda.fill = PatternFill("solid", fgColor=GRIS)

        # Las columnas que llena el inspector van en amarillo, en ambas hojas.
        for clave in ("decision_final", "obs_inspector"):
            col = next((i for i, c in enumerate(columnas, 1) if c[2] == clave), None)
            if col is not None:
                ws.cell(row=n, column=col).fill = PatternFill("solid", fgColor=AMARILLO)


def _fila_sin_hallazgos(fecha, categoria):
    """Fila testigo: deja constancia de que la categoría SÍ se revisó ese día."""
    return {
        "nombre_dm": "Sin hallazgos",
        "url": "",
        "titulo": "",
        "oferente": "",
        "coincidencia": "",
        "producto_isp": "",
        "registro_isp": "",
        "clasificacion": "SIN HALLAZGOS",
        "observaciones": (
            f"Revisión de la categoría «{categoria}» realizada el {fecha} sin detectar "
            "publicaciones individuales de productos no registrados. La ausencia de "
            "hallazgos no descarta ofertas no indexadas por los buscadores ni las "
            "alojadas en sitios con acceso restringido."
        ),
        "decision_final": "",
        "obs_inspector": "",
    }


def validar(datos):
    """Controles de calidad sobre los hallazgos, antes de emitir el Excel.

    Devuelve una lista de avisos. No corrige nada ni descarta filas: la decisión
    es del fiscalizador. Solo hace visible lo que de otro modo llega al inspector
    sin que nadie lo note.
    """
    avisos = []
    hallazgos = datos.get("hallazgos") or []
    if not hallazgos:
        return avisos

    categoria = str(datos.get("categoria", "")).strip().lower()
    nombres = [str(h.get("nombre_dm", "")).strip() for h in hallazgos]

    # La columna 1 debe llevar el nombre publicado del producto, no la categoría.
    # Si todas las filas dicen lo mismo, el reporte no le sirve al inspector:
    # no puede distinguir un producto de otro.
    genericos = [n for n in nombres if n.lower() == categoria]
    if genericos:
        avisos.append(
            f"{len(genericos)} de {len(nombres)} filas usan el nombre de la categoría "
            f"(«{datos.get('categoria')}») como nombre del producto. La columna "
            "'Nombre de DM ofertado' debe llevar el nombre tal como aparece publicado."
        )
    elif len(set(n.lower() for n in nombres if n)) == 1 and len(nombres) > 1:
        avisos.append(
            f"Las {len(nombres)} filas repiten el mismo nombre de producto "
            f"(«{nombres[0]}»). Debe ir el nombre publicado de cada oferta."
        )

    # Un REGISTRADO sin número de registro no es verificable por el inspector.
    sin_reg = [
        h.get("nombre_dm") for h in hallazgos
        if str(h.get("clasificacion", "")).strip().upper() == "REGISTRADO"
        and not str(h.get("registro_isp", "")).strip()
    ]
    if sin_reg:
        avisos.append(
            f"{len(sin_reg)} hallazgo(s) marcados REGISTRADO sin N° de registro sanitario. "
            "Sin ese dato la clasificación no es verificable."
        )

    urls = [str(h.get("url", "")).strip() for h in hallazgos if h.get("url")]
    rep = {u for u in urls if urls.count(u) > 1}
    if rep:
        avisos.append(f"{len(rep)} URL(s) repetidas en el reporte: {', '.join(list(rep)[:3])}")

    # Guantes: la regulación cubre solo examinación y quirúrgicos DE LÁTEX. Un
    # guante de otro material, o de látex pero de uso no médico, no es una
    # infracción aunque no tenga registro. Se avisa en vez de excluir: la
    # decisión es del fiscalizador, pero no puede pasar inadvertida.
    if "guante" in _plano(categoria):
        acusados = [
            h for h in hallazgos
            if str(h.get("clasificacion", "")).strip().upper() == "NO REGISTRADO"
        ]
        otro_material = [
            h for h in acusados
            if any(m in _plano(f"{h.get('nombre_dm','')} {h.get('titulo','')} "
                               f"{h.get('observaciones','')}")
                   for m in MATERIALES_FUERA_ALCANCE)
        ]
        if otro_material:
            avisos.append(
                f"{len(otro_material)} hallazgo(s) acusados como NO REGISTRADO mencionan un "
                "material distinto del látex (nitrilo, vinilo u otro). La regulación cubre "
                "SOLO guantes de examinación y quirúrgicos de látex: esos productos están "
                "FUERA DE ALCANCE y no son infracción. Ejemplo: "
                f"«{otro_material[0].get('nombre_dm','')}»."
            )
        # Solo los que no salieron ya por material: si no, la misma fila se
        # cuenta dos veces y el número del aviso engaña.
        ids_material = {id(h) for h in otro_material}
        sin_uso = [
            h for h in acusados
            if id(h) not in ids_material
            and not any(u in _plano(f"{h.get('nombre_dm','')} {h.get('titulo','')}")
                        for u in USO_MEDICO)
        ]
        if sin_uso:
            avisos.append(
                f"{len(sin_uso)} hallazgo(s) acusados como NO REGISTRADO no declaran uso de "
                "examinación ni quirúrgico en el nombre ni en el título. Verifica en la "
                "publicación que sean de uso médico: un guante de látex para tatuaje, "
                "cosmetología o aseo está FUERA DE ALCANCE. Ejemplo: "
                f"«{sin_uso[0].get('nombre_dm','')}»."
            )

    marcadores = ("listado.", "/search", "?q=", "/buscar", "/s?", "google.com/search")
    busq = [u for u in urls if any(m in u.lower() for m in marcadores)]
    if busq:
        avisos.append(
            f"{len(busq)} URL(s) parecen páginas de búsqueda, no publicaciones "
            f"individuales: {busq[0]}"
        )

    return avisos


def priorizar(hallazgos):
    """Ordena por valor para el inspector y separa lo que excede el tope.

    Un NO REGISTRADO es un caso que hay que investigar; un REGISTRADO es una
    confirmación. Si hay que recortar, se recorta por el lado que menos trabajo
    genera, nunca dejando fuera una posible infracción para dejar dentro un
    producto que ya sabemos que cumple.
    """
    def rango(h):
        c = str(h.get("clasificacion", "")).strip().upper()
        return {"NO REGISTRADO": 0, "REGISTRADO": 1}.get(c, 2)

    ordenados = sorted(hallazgos, key=rango)
    return ordenados[:MAX_HALLAZGOS], ordenados[MAX_HALLAZGOS:]


def generar(datos, salida):
    categoria = datos.get("categoria", "(sin categoría)")
    fecha = datos.get("fecha") or date.today().strftime("%d-%m-%Y")
    todos = datos.get("hallazgos") or []
    hallazgos, excedentes = priorizar(todos)
    marketplace = datos.get("marketplace") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Hallazgos"
    _encabezados(ws, COLUMNAS)

    filas = hallazgos if hallazgos else [_fila_sin_hallazgos(fecha, categoria)]
    _escribir_filas(ws, COLUMNAS, filas)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{len(filas) + 1}"

    if excedentes:
        ws3 = wb.create_sheet("Anexo — sobre el tope")
        _encabezados(ws3, COLUMNAS_ANEXO)
        filas_anexo = [
            {
                "nombre_dm": h.get("nombre_dm", ""),
                "url": h.get("url", ""),
                "oferente": h.get("oferente", ""),
                "clasificacion": h.get("clasificacion", ""),
                "observaciones": h.get("observaciones", ""),
            }
            for h in excedentes
        ]
        _escribir_filas(ws3, COLUMNAS_ANEXO, filas_anexo, colorear_clasificacion=False)
        ws3.cell(
            row=len(filas_anexo) + 2,
            column=1,
            value=(
                f"Estas {len(excedentes)} ofertas se detectaron en la misma búsqueda pero "
                f"quedaron fuera del reporte por el tope de {MAX_HALLAZGOS} hallazgos. "
                "NO requieren revisión esta semana. Al no quedar registradas como "
                "evaluadas, vuelven a considerarse en la próxima corrida de la categoría."
            ),
        ).font = Font(italic=True, size=9)

    if marketplace:
        ws2 = wb.create_sheet("Búsquedas Marketplace")
        _encabezados(ws2, COLUMNAS_MKT)
        _escribir_filas(ws2, COLUMNAS_MKT, marketplace, colorear_clasificacion=False)

    os.makedirs(os.path.dirname(os.path.abspath(salida)), exist_ok=True)
    wb.save(salida)

    return {
        "archivo": salida,
        "total": len(hallazgos),
        "total_detectado": len(todos),
        "excedentes": len(excedentes),
        "tope": MAX_HALLAZGOS,
        "no_registrado": sum(
            1 for h in hallazgos
            if str(h.get("clasificacion", "")).strip().upper() == "NO REGISTRADO"
        ),
        "registrado": sum(
            1 for h in hallazgos
            if str(h.get("clasificacion", "")).strip().upper() == "REGISTRADO"
        ),
        "marketplace": len(marketplace),
        "sin_hallazgos": not hallazgos,
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--entrada", required=True, help="JSON con los hallazgos")
    ap.add_argument("--salida", help="ruta del .xlsx a generar")
    ap.add_argument(
        "--auto",
        action="store_true",
        help="deducir la ruta de salida desde scripts/rotacion.py",
    )
    ap.add_argument("--slot", type=int, choices=(1, 2), default=1,
                    help="bloque del día, usado junto con --auto")
    ap.add_argument("--fecha", help="fecha YYYY-MM-DD, usada junto con --auto")
    args = ap.parse_args()

    with open(args.entrada, encoding="utf-8") as fh:
        datos = json.load(fh)

    salida = args.salida
    if not salida:
        if not args.auto:
            print("Falta --salida (o usa --auto).", file=sys.stderr)
            return 1
        sys.path.insert(0, os.path.join(RAIZ, "scripts"))
        import rotacion
        from datetime import date as _date
        fecha = _date.fromisoformat(args.fecha) if args.fecha else None
        plan = rotacion.construir_plan(fecha, args.slot)
        if not plan.get("habil"):
            print(f"{plan['dia']} {plan['fecha']}: {plan['motivo']}", file=sys.stderr)
            return 3
        salida = plan["archivo_salida"]
        datos.setdefault("categoria", plan["categoria"])
        datos.setdefault("fecha", plan["fecha"])

    avisos = validar(datos)
    resumen = generar(datos, salida)
    if avisos:
        resumen["avisos_calidad"] = avisos

    # El reporte va topado, pero el resumen informa sobre TODO lo revisado: la
    # mezcla registrado/no registrado sigue siendo la señal de si la búsqueda
    # fue profunda o se quedó en las tiendas grandes.
    detectado = resumen["total_detectado"]
    if detectado:
        nr = sum(1 for h in (datos.get("hallazgos") or [])
                 if str(h.get("clasificacion", "")).strip().upper() == "NO REGISTRADO")
        rg = sum(1 for h in (datos.get("hallazgos") or [])
                 if str(h.get("clasificacion", "")).strip().upper() == "REGISTRADO")
        resumen["detectado_no_registrado"] = nr
        resumen["detectado_registrado"] = rg
        resumen["pct_no_registrado_detectado"] = round(100 * nr / detectado, 1)
        if rg == detectado:
            resumen["aviso"] = (
                "TODOS los hallazgos salieron REGISTRADO. Señal de que la búsqueda se "
                "quedó en las tiendas grandes y formales. Haz otra ronda antes de cerrar."
            )
    if resumen["excedentes"]:
        resumen["nota_tope"] = (
            f"Se detectaron {detectado} ofertas; se incluyen las {resumen['total']} más "
            f"relevantes por el tope de {resumen['tope']}. Las {resumen['excedentes']} "
            "restantes quedaron en la hoja «Anexo» y reaparecerán en la próxima corrida."
        )

    print(json.dumps(resumen, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
