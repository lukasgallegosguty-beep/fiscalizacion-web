#!/usr/bin/env python3
"""
Lee los Excel ya revisados por los inspectores y los convierte en insumo para la
búsqueda siguiente.

Resuelve tres cosas que el prompt por sí solo no garantiza:

1. **No repetir enlaces.** Junta todas las URL que un inspector ya evaluó, para
   excluirlas de los hallazgos nuevos.
2. **Corregir el criterio de cruce.** Compara la clasificación automática contra
   la "Decisión final" del inspector y separa falsos positivos de falsos
   negativos.
3. **Incorporar las indicaciones.** Recoge el texto libre de "Observaciones del
   inspector", en ambas hojas, como instrucciones para la corrida.

Uso:
    python3 scripts/feedback.py --categoria agujas-hipodermicas
    python3 scripts/feedback.py --categoria agujas-hipodermicas --json
    python3 scripts/feedback.py --todas
"""

import argparse
import glob
import json
import os
import sys

from openpyxl import load_workbook

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DIR_REVISION = os.path.join(RAIZ, "revision")
DIR_RESULTADOS = os.path.join(RAIZ, "resultados")


def _norm(v):
    if v is None:
        return ""
    return (
        str(v).strip().lower()
        .replace("á", "a").replace("é", "e").replace("í", "i")
        .replace("ó", "o").replace("ú", "u").replace("ñ", "n")
    )


def _veredicto(texto):
    """Interpreta la 'Decisión final'. Devuelve REGISTRADO, NO REGISTRADO u OTRO."""
    t = _norm(texto)
    if not t:
        return None
    if "no registrado" in t or "sin registro" in t:
        return "NO REGISTRADO"
    if "registrado" in t or "si tiene registro" in t:
        return "REGISTRADO"
    return "OTRO"


def _indice(ws):
    """Mapea nombre de encabezado -> número de columna, desde la fila 1."""
    idx = {}
    for celda in ws[1]:
        if celda.value:
            idx[_norm(celda.value)] = celda.column
    return idx


def _leer_archivo(ruta):
    wb = load_workbook(ruta, read_only=True, data_only=True)
    datos = {"archivo": os.path.basename(ruta), "filas": [], "marketplace": []}

    if "Hallazgos" in wb.sheetnames:
        ws = wb["Hallazgos"]
        idx = _indice(ws)
        c_url = idx.get("url")
        c_dec = idx.get("decision final")
        c_obs = idx.get("observaciones del inspector")
        c_cla = idx.get("clasificacion")
        c_nom = idx.get("nombre de dm ofertado")
        for fila in ws.iter_rows(min_row=2):
            def val(c):
                return fila[c - 1].value if c and len(fila) >= c else None
            url = (val(c_url) or "").strip() if isinstance(val(c_url), str) else val(c_url)
            datos["filas"].append({
                "nombre": val(c_nom),
                "url": url or "",
                "clasificacion_auto": (val(c_cla) or "").strip() if isinstance(val(c_cla), str) else "",
                "decision_final": val(c_dec) or "",
                "obs_inspector": val(c_obs) or "",
            })

    for hoja in wb.sheetnames:
        if hoja.startswith("Búsquedas"):
            ws = wb[hoja]
            idx = _indice(ws)
            c_obs = idx.get("observaciones del inspector")
            c_mkt = idx.get("marketplace")
            for fila in ws.iter_rows(min_row=2):
                def val(c):
                    return fila[c - 1].value if c and len(fila) >= c else None
                obs = val(c_obs)
                if obs and str(obs).strip():
                    datos["marketplace"].append({
                        "marketplace": val(c_mkt), "obs_inspector": str(obs).strip()
                    })
    wb.close()
    return datos


def analizar(slug):
    """Consolida el feedback de todos los archivos revisados de una categoría."""
    patron = f"Fiscalizacion_Web_DM_{slug}_*.xlsx"
    archivos = sorted(glob.glob(os.path.join(DIR_REVISION, patron)), key=os.path.getmtime)

    res = {
        "categoria": slug,
        "archivos_revisados": [],
        "urls_excluidas": [],
        "falsos_positivos": [],
        "falsos_negativos": [],
        "confirmados": 0,
        "instrucciones_inspector": [],
        "obs_marketplace": [],
    }

    vistas = set()
    for ruta in archivos:
        d = _leer_archivo(ruta)
        res["archivos_revisados"].append(d["archivo"])

        for f in d["filas"]:
            url = (f["url"] or "").strip()
            # Toda URL que el inspector ya vio queda fuera de las corridas siguientes.
            if url and url not in vistas:
                vistas.add(url)
                res["urls_excluidas"].append(url)

            ver = _veredicto(f["decision_final"])
            auto = (f["clasificacion_auto"] or "").strip().upper()
            if ver in ("REGISTRADO", "NO REGISTRADO") and auto in ("REGISTRADO", "NO REGISTRADO"):
                if ver == auto:
                    res["confirmados"] += 1
                elif auto == "NO REGISTRADO" and ver == "REGISTRADO":
                    # La rutina lo acusó y el inspector lo desmintió: criterio muy estricto.
                    res["falsos_positivos"].append(
                        {"nombre": f["nombre"], "url": url, "decision": str(f["decision_final"]).strip()}
                    )
                else:
                    # La rutina lo dio por registrado y sí era infracción: criterio muy laxo.
                    res["falsos_negativos"].append(
                        {"nombre": f["nombre"], "url": url, "decision": str(f["decision_final"]).strip()}
                    )

            obs = str(f["obs_inspector"]).strip()
            if obs:
                res["instrucciones_inspector"].append(
                    {"origen": d["archivo"], "sobre": f["nombre"], "texto": obs}
                )

        for m in d["marketplace"]:
            res["obs_marketplace"].append({"origen": d["archivo"], **m})

    res["total_urls_excluidas"] = len(res["urls_excluidas"])
    res["hay_feedback"] = bool(archivos)
    return res


def imprimir(res):
    if not res["hay_feedback"]:
        print(f"[{res['categoria']}] sin archivos revisados en revision/ — primera corrida de la categoría")
        return
    print(f"[{res['categoria']}] {len(res['archivos_revisados'])} archivo(s) revisado(s)")
    print(f"  URLs ya evaluadas (excluir) : {res['total_urls_excluidas']}")
    print(f"  Clasificaciones confirmadas : {res['confirmados']}")
    print(f"  Falsos positivos            : {len(res['falsos_positivos'])}  -> criterio demasiado estricto")
    print(f"  Falsos negativos            : {len(res['falsos_negativos'])}  -> criterio demasiado laxo")
    if res["instrucciones_inspector"]:
        print(f"\n  Indicaciones del inspector ({len(res['instrucciones_inspector'])}):")
        for i in res["instrucciones_inspector"]:
            print(f"    - ({i['sobre']}) {i['texto']}")
    if res["obs_marketplace"]:
        print(f"\n  Observaciones en marketplace ({len(res['obs_marketplace'])}):")
        for m in res["obs_marketplace"]:
            print(f"    - [{m['marketplace']}] {m['obs_inspector']}")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--categoria", help="slug de la categoría")
    ap.add_argument("--todas", action="store_true", help="todas las categorías")
    ap.add_argument("--json", action="store_true", help="salida JSON")
    args = ap.parse_args()

    if args.todas:
        sys.path.insert(0, os.path.join(RAIZ, "scripts"))
        import rotacion
        todo = [analizar(c["slug"]) for c in rotacion.CATEGORIAS]
        if args.json:
            print(json.dumps(todo, ensure_ascii=False, indent=2))
        else:
            for r in todo:
                imprimir(r)
        return 0

    if not args.categoria:
        print("Falta --categoria (o usa --todas).", file=sys.stderr)
        return 1

    res = analizar(args.categoria)
    if args.json:
        print(json.dumps(res, ensure_ascii=False, indent=2))
    else:
        imprimir(res)
    return 0


if __name__ == "__main__":
    sys.exit(main())
